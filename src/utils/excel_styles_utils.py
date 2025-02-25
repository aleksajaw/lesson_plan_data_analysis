from error_utils import handleErrorMsg, getTraceback
#from src.constants.paths_constants import scheduleClassesExcelPath
from src.constants.excel_constants import excelMargin, excelFontSize, excelRangeStartCol, excelRangeStartRow
from src.constants.schedule_structures_constants import defRowNamesLen, defColNamesLen#, timeIndexNames, dayAndAttrNames
#from pandas import ExcelWriter
#from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles import Alignment as openpyxlAlignment
from openpyxl.styles import PatternFill as openpyxlPatternFill
from openpyxl.styles import Border as openpyxlBorder
from openpyxl.styles import Side as openpyxlSide
from openpyxl.styles import Font as openpyxlFont
#from openpyxl.cell.cell import Cell as openpyxlCell
#from openpyxl.cell.cell import MergedCell as openpyxlMergedCell
from openpyxl.utils import get_column_letter
#from openpyxl.worksheet.worksheet import Worksheet



def autoFormatScheduleExcel(workbook, doesNeedFormatStyle=True,  dfRowIndexLen=defRowNamesLen, dfColNamesLen=defColNamesLen, doesNeedFormatSize=True):
    if doesNeedFormatSize:
        autoFormatExcelCellSizes(workbook)
    if doesNeedFormatStyle:
        autoFormatScheduleExcelCellStyles(workbook, dfRowIndexLen, dfColNamesLen)



def autoFormatOverviewExcel(workbook, doesNeedFormatStyle=False):
    #if workbook is None:
    #    workbook = Workbook
    autoFormatExcelCellSizes(workbook)
    if doesNeedFormatStyle:
        autoFormatOverviewExcelCellStyles(workbook, dfRowIndexLen=3)



def autoFormatOverviewExcelCellStyles(workbook, dfRowIndexLen=defRowNamesLen, dfColNamesLen=defColNamesLen, shouldPrintSuccessMsg=False):
    from src.utils.excel_utils import getNrOfLastNonEmptyCellInCol
    msgText=''

    try:
        #if workbook is None:
        #    workbook = Workbook
        
        for ws in workbook.worksheets:
            
            # column nr where the row indices end
            rowIndexesLastCol = excelMargin['col'] + dfRowIndexLen
            rowNumericIndexCol = rowIndexesLastCol-1

            # bold rows are for headers
            #lastBoldRowAtBeggining = findLastBoldRowAtBeggining(ws, rowIndexesLastCol, excelRangeStartRow)
            lastBoldRowAtBeggining = excelMargin['row'] + dfColNamesLen

            # merge and format empty cells in the corner between the MultiIndexes
            mergeEmptyCellsAndFormat(ws, { 'startRow':  excelRangeStartRow,
                                            'startCol':  excelRangeStartCol,
                                            'endRow'  :  lastBoldRowAtBeggining,
                                            'endCol'  :  rowIndexesLastCol } )
            
            # cells in the 1st two columns which row nr equals contentRowsStart
            # contains the names for the rows' MultiIndex
            firstRowUnderColNames = lastBoldRowAtBeggining + ( 1   if dfRowIndexLen   else 0)
            # col nr for the 1st cell for the 1st day
            daysFirstColList = rowIndexesLastCol+1

            totalColsRange = range(excelRangeStartCol, ws.max_column+1)
            totalColsRangeToColor = range(rowNumericIndexCol, ws.max_column+1)
            # cells in the 2nd column contain part of rows indices & are less likely to be merged (for easier data analysis)
            totalRowsCount = lastBoldRowAtBeggining + getNrOfLastNonEmptyCellInCol(ws, minRow=firstRowUnderColNames, col=rowIndexesLastCol)
            totalRowsRange = range(excelRangeStartRow, totalRowsCount+1)


            colorBgOfEmptyRow(ws, range(rowIndexesLastCol+1, ws.max_column+1), row=firstRowUnderColNames, startColumn=daysFirstColList)


            # add MEDIUM RIGHT BORDERS
            # at the end of each important column and each day
            daysLastCol = sorted([mergedCellRange.max_col   for mergedCellRange in ws.merged_cells
                                                                if mergedCellRange.min_row == mergedCellRange.max_row == excelRangeStartRow
                                                                  and   mergedCellRange.min_col > rowIndexesLastCol])
            
            rowIndexesCols = list( range(excelRangeStartCol, rowIndexesLastCol+1) )
            colsWithRightMediumBorder = rowIndexesCols + daysLastCol

            if excelMargin['col']:
                colsWithRightMediumBorder.insert(0, excelMargin['col'])
            
            standardDaySize = daysLastCol[0] - rowIndexesLastCol

            for col in colsWithRightMediumBorder:

                for row1 in totalRowsRange:
                    cell = ws.cell(row=row1, column=col)
                    formatCellBorder(cell, right='medium')
                    
                    # add THIN SIDE BORDERS to the center column on days
                    # do not include the columns reserved for the rows indices
                    if rowIndexesLastCol < col:
                        rangeStartTemp = (col-standardDaySize)+1

                        for attrCol in range(rangeStartTemp, col):
                            cell = ws.cell(row=row1, column=attrCol)
                            formatCellBorder(cell, left='thin', right='thin')
            


            # add MEDIUM BOTTOM BORDERS
            # at the end of each index (for columns, for rows) and the entire schedule
            rowsWithBottomBorder = [lastBoldRowAtBeggining, firstRowUnderColNames, totalRowsCount]

            if excelMargin['row']:
                rowsWithBottomBorder.insert(0, excelMargin['row'])

            # Add a bottom border at the end of each day.
            if excelRangeStartCol != rowNumericIndexCol:
                mergedCellsFirstColMaxRow = sorted( [ r.max_row   for r in ws.merged_cells.ranges
                                                                      if r.min_col == excelRangeStartCol
                                                                        and   r.min_row > firstRowUnderColNames
                                                                        and   r.max_row < totalRowsCount ] )

                rowsWithBottomBorder = rowsWithBottomBorder[:-1] + mergedCellsFirstColMaxRow + rowsWithBottomBorder[-1:]

            for row in rowsWithBottomBorder:
                
                # ignore long merged cell
                totalColsRangeTemp = ( totalColsRange   if row != firstRowUnderColNames
                                                        else range(excelRangeStartCol, rowIndexesLastCol+1) )

                for col in totalColsRangeTemp:
                    cell = ws.cell(row, column=col)
                    formatCellBorder(cell, bottom='medium')



            # add BACKGROUND
            # for the lessons with an odd numbers
            # and HAIR/THIN BORDERS
            for row in totalRowsRange:

                for col in totalColsRange:
                    cell = ws.cell(row, column=col)
                    # thinner ('HAIR') BORDER inside merged row
                    #if merged   and   merged.min_row != row:
                    #    topBorderStyle = 'hair'
                    # THIN BORDER for outer frame
                    #elif not cell.border.top.style:
                    topBorderStyle = 'thin'
                    formatCellBorder(cell, top=topBorderStyle)

                if row > lastBoldRowAtBeggining:
                    # fast checking if the row is in the specific merged range (for col A here)
                    merged = next(  ( r   for r in ws.merged_cells.ranges
                                              if r.min_col == rowNumericIndexCol
                                                and   r.min_row <= row <= r.max_row ),
                                    None)
                    


                    # set LIGHT GREY BACKGROUND
                    # for the lessons with an odd number
                    rowNr = merged.min_row   if merged   else row
                    cellValue = ws.cell(row=rowNr, column=rowNumericIndexCol).value

                    if isinstance(cellValue, int)   and   (cellValue & 1):
                        
                        for col in totalColsRangeToColor:
                            cell = ws.cell(row, column=col)
                            formatCellBackground(cell, fillType='solid', startColor='E5E5E5', endColor='E5E5E5')


        if shouldPrintSuccessMsg: msgText = '\nThe cell styles of the Excel file auto formatted.'

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell styles in the Excel file.', getTraceback(e))
      
    if msgText: print(msgText)



def autoFormatScheduleExcelCellStyles(workbook, dfRowIndexLen=defRowNamesLen, dfColNamesLen=defColNamesLen, shouldPrintSuccessMsg=False):
    from src.utils.excel_utils import getNrOfLastNonEmptyCellInCol
    msgText=''

    try:
        #if workbook is None:
        #    workbook = Workbook
        
        for ws in workbook.worksheets:
            
            # column nr where the row indices end
            rowIndexesLastCol = excelMargin['col'] + dfRowIndexLen
            rowNumericIndexCol = rowIndexesLastCol-1

            # bold rows are for headers
            #lastBoldRowAtBeggining = findLastBoldRowAtBeggining(ws, rowIndexesLastCol, excelRangeStartRow)
            lastBoldRowAtBeggining = excelMargin['row'] + dfColNamesLen

            # merge and format empty cells in the corner between the MultiIndexes
            mergeEmptyCellsAndFormat(ws, { 'startRow':  excelRangeStartRow,
                                            'startCol':  excelRangeStartCol,
                                            'endRow'  :  lastBoldRowAtBeggining,
                                            'endCol'  :  rowIndexesLastCol } )
            
            # cells in the 1st two columns which row nr equals contentRowsStart
            # contains the names for the rows' MultiIndex
            firstRowUnderColNames = lastBoldRowAtBeggining + ( 1   if dfRowIndexLen   else 0)
            # col nr for the 1st cell for the 1st day
            daysFirstColList = rowIndexesLastCol+1

            totalColsRange = range(excelRangeStartCol, ws.max_column+1)
            totalColsRangeToColor = range(rowNumericIndexCol, ws.max_column+1)
            # cells in the 2nd column contain part of rows indices & are less likely to be merged (for easier data analysis)
            totalRowsCount = lastBoldRowAtBeggining + getNrOfLastNonEmptyCellInCol(ws, minRow=firstRowUnderColNames, col=rowIndexesLastCol)
            totalRowsRange = range(excelRangeStartRow, totalRowsCount+1)


            colorBgOfEmptyRow(ws, range(rowIndexesLastCol+1, ws.max_column+1), row=firstRowUnderColNames, startColumn=daysFirstColList)


            # add MEDIUM RIGHT BORDERS
            # at the end of each important column and each day
            daysLastCol = sorted([mergedCellRange.max_col   for mergedCellRange in ws.merged_cells
                                                                if mergedCellRange.min_row == mergedCellRange.max_row == excelRangeStartRow
                                                                  and   mergedCellRange.min_col > rowIndexesLastCol])
            
            rowIndexesCols = list( range(excelRangeStartCol, rowIndexesLastCol+1) )
            colsWithRightMediumBorder = rowIndexesCols + daysLastCol

            if excelMargin['col']:
                colsWithRightMediumBorder.insert(0, excelMargin['col'])
            
            standardDaySize = daysLastCol[0] - rowIndexesLastCol

            for col in colsWithRightMediumBorder:

                for row1 in totalRowsRange:
                    cell = ws.cell(row=row1, column=col)
                    formatCellBorder(cell, right='medium')
                    
                    # add THIN SIDE BORDERS to the center column on days
                    # do not include the columns reserved for the rows indices
                    if rowIndexesLastCol < col:
                        rangeStartTemp = (col-standardDaySize)+1

                        for attrCol in range(rangeStartTemp, col):
                            cell = ws.cell(row=row1, column=attrCol)
                            formatCellBorder(cell, left='thin', right='thin')
            


            # add MEDIUM BOTTOM BORDERS
            # at the end of each index (for columns, for rows) and the entire schedule
            rowsWithBottomBorder = [lastBoldRowAtBeggining, firstRowUnderColNames, totalRowsCount]

            if excelMargin['row']:
                rowsWithBottomBorder.insert(0, excelMargin['row'])

            # Add a bottom border at the end of each day.
            if excelRangeStartCol != rowNumericIndexCol:
                mergedCellsFirstColMaxRow = sorted( [ r.max_row   for r in ws.merged_cells.ranges
                                                                      if r.min_col == excelRangeStartCol
                                                                        and   r.min_row > firstRowUnderColNames
                                                                        and   r.max_row < totalRowsCount ] )

                rowsWithBottomBorder = rowsWithBottomBorder[:-1] + mergedCellsFirstColMaxRow + rowsWithBottomBorder[-1:]

            for row in rowsWithBottomBorder:
                
                # ignore long merged cell
                totalColsRangeTemp = ( totalColsRange   if row != firstRowUnderColNames
                                                        else range(excelRangeStartCol, rowIndexesLastCol+1) )

                for col in totalColsRangeTemp:
                    cell = ws.cell(row, column=col)
                    formatCellBorder(cell, bottom='medium')
                    cell2 = ws.cell(row=row+1, column=col)
                    formatCellBorder(cell2, top='medium')



            # add BACKGROUND
            # for the lessons with an odd numbers
            # and HAIR/THIN BORDERS
            for row in totalRowsRange:

                for col in totalColsRange:
                    cell = ws.cell(row, column=col)
                    # thinner ('HAIR') BORDER inside merged row
                    #if merged   and   merged.min_row != row:
                    #    topBorderStyle = 'hair'
                    # THIN BORDER for outer frame
                    #elif not cell.border.top.style:
                    topBorderStyle = 'thin'
                    formatCellBorder(cell, top=topBorderStyle)

                if row > lastBoldRowAtBeggining:
                    # fast checking if the row is in the specific merged range (for col A here)
                    merged = next(  ( r   for r in ws.merged_cells.ranges
                                              if r.min_col == rowNumericIndexCol
                                                and   r.min_row <= row <= r.max_row ),
                                    None)
                    


                    # set LIGHT GREY BACKGROUND
                    # for the lessons with an odd number
                    rowNr = merged.min_row   if merged   else row
                    cellValue = ws.cell(row=rowNr, column=rowNumericIndexCol).value

                    if isinstance(cellValue, int)   and   (cellValue & 1):
                        
                        for col in totalColsRangeToColor:
                            cell = ws.cell(row, column=col)
                            formatCellBackground(cell, fillType='solid', startColor='E5E5E5', endColor='E5E5E5')


        if shouldPrintSuccessMsg: msgText = '\nThe cell styles of the Excel file auto formatted.'

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell styles in the Excel file.', getTraceback(e))
      
    if msgText: print(msgText)


    
def autoFormatExcelCellSizes(workbook, shouldPrintSuccessMsg=False):
    msgText = ''
    
    try:
        #if workbook is None:
        #    workbook = Workbook
        
        for ws in workbook.worksheets:
                
            for col in range(excelRangeStartCol, ws.max_column+1):
                max_length = 0
                colLetter = get_column_letter(col)
                
                # loop through cells in column
                for cell in ws[colLetter]:
                    isCellMergedHorizontally = any(cell.coordinate in merged_range   and   merged_range.min_col != merged_range.max_col   for merged_range in ws.merged_cells.ranges)
                    
                    if cell.row > excelMargin['row']   and   not isCellMergedHorizontally:
                        
                        try:
                            cellValueLen = len(str(cell.value))

                            if cellValueLen > max_length:
                                max_length = cellValueLen
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[colLetter].width = adjusted_width

            for row in range(excelRangeStartRow, ws.max_row+1):
                
                for cell in ws[row]:
                    
                    cell.alignment = openpyxlAlignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.font = openpyxlFont(size=excelFontSize, bold=cell.font.bold)
                
                ws.row_dimensions[row].height = excelFontSize * 1.3


        if shouldPrintSuccessMsg: msgText = '\nThe cell sizes of the Excel file auto formatted.'

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell sizes in the Excel file.', getTraceback(e))

    if msgText: print(msgText)



def formatCellBorder(cell, right='', left='', top='', bottom=''):
    msgText = ''

    try:
        #if cell is None:
        #    cell = openpyxlCell #| openpyxlMergedCell)
        
        currentBorder = cell.border
        defaultColor = '000000'
        borderStyle = { 'hair'  :  openpyxlSide(border_style='hair',    color=defaultColor),
                        'thin'  :  openpyxlSide(border_style='thin',    color=defaultColor),
                        'medium':  openpyxlSide(border_style='medium',  color=defaultColor),
                        'thick' :  openpyxlSide(border_style='thick',   color=defaultColor) }
        
        cell.border = openpyxlBorder( right = borderStyle[right]     if right    else currentBorder.right,
                                      left = borderStyle[left]       if left     else currentBorder.left,
                                      top = borderStyle[top]         if top      else currentBorder.top,
                                      bottom = borderStyle[bottom]   if bottom   else currentBorder.bottom )

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell.', getTraceback(e))
    
    if msgText: print(msgText)



def formatCellBackground(cell, fillType, startColor='', endColor=''):
    msgText = ''

    try:
        #if cell is None:
        #    cell = openpyxlCell #| openpyxlMergedCell)
        
        if ( fillType=='solid'
                and   isinstance(startColor, str)   and   startColor
                and   isinstance(endColor, str)     and   endColor ):
            
            cell.fill = openpyxlPatternFill(start_color=startColor, end_color=endColor, fill_type=fillType)
        
        elif fillType:
            cell.fill = openpyxlPatternFill(fill_type=fillType)

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell.', getTraceback(e))

    if msgText: print(msgText)



def addBgToExcelSheetRowsBasedOnObj(workbook, sheetRowsToColor={'rows':[], 'colsLength':0}):
    # add BACKGROUND to the (odd here) groups of the cells in worksheet 
    msgText = ''

    try:
        #if workbook is None:
        #    workbook = Workbook
        
        for sheetname in workbook.sheetnames:
            ws = workbook[sheetname]
            sheetBgRanges = sheetRowsToColor[sheetname]

            for grRow in sheetBgRanges['rows']:
                
                for col in range(excelRangeStartCol, ws.max_column+1):
                  
                    cell = ws.cell(row=grRow, column=col)
                    formatCellBackground(cell, 'solid', 'f3f3f3', 'f3f3f3')
    
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while adding background to the cells in the Excel sheet rows.', getTraceback(e))
    
    if msgText: print(msgText)


def mergeEmptyCellsAndFormat(ws, mergedCellObj={'startRow':1, 'startCol':1, 'endRow':1, 'endCol':1}):
    msgText=''

    try:
        #if ws is None:
        #    ws = Worksheet
        
        permEmptyCellStyle = openpyxlPatternFill(fill_type='lightTrellis')

        startRow = mergedCellObj['startRow']
        startCol = mergedCellObj['startCol']
        endRow = mergedCellObj['endRow']
        endCol = mergedCellObj['endCol']

        endColTemp = -1
        endRowTemp = -1
        
        for col in range(endCol, startCol-1, -1):
            cell = None

            for row in range(endRow, startRow-1, -1):
                cell = ws.cell(row=row, column=col)
                if cell:
                    if cell.value:
                        endRowTemp = -1
                        endColTemp = -1

                    else:
                        if endRowTemp!=row+1:
                            endRowTemp = row
                        if endColTemp!=col+1:
                            endColTemp = col
            
        endCol = endColTemp
        endRow = endRowTemp

        if ( startRow>0   and   startCol>0):
            if ( endRow>=startRow   and   endCol>=startCol   and 
               ( endRow>startRow    or   endCol>startCol) ):
                
                ws.merge_cells( start_row=startRow,
                                start_column=startCol,
                                end_row=endRow,
                                end_column=endCol )
            
            cell = ws.cell(row=startRow, column=startCol)
            
            if not cell.value:
                cell.fill = permEmptyCellStyle
                #formatCellBorder(cell, right='medium', bottom='medium')


    except Exception as e:
        msgText = handleErrorMsg('\nError while merging the empty cells and coloring their background.', getTraceback(e))


    if msgText: print(msgText)



def colorBgOfEmptyRow(ws, colRange, row=0, startColumn=0):
    msgText=''

    try:
        #if ws is None:
        #    ws = Worksheet
        #if colRange is None:
        #    colRange = range(col)
        
        permEmptyCellStyle = openpyxlPatternFill(fill_type='lightTrellis')
        allEmpty = True 

        # check if the cells in the 1st row (below column indices) are empty
        for col in colRange:
            
            if ws.cell(row=row, column=col).value is not None:
                allEmpty = False
                break


        # change the BACKGROUND of the cells in the row (below column indices)
        # only if entire row is empty
        if allEmpty:
            
            # MERGE CELLS in the row 
            ws.merge_cells( start_row=row, start_column=startColumn,
                            end_row=row,   end_column=max(colRange) )
            
            cell = ws.cell(row=row, column=startColumn)
            cell.fill = permEmptyCellStyle
            

    except Exception as e:
        msgText = handleErrorMsg('\nError while coloring the background of empty rows.', getTraceback(e))

    if msgText: print(msgText)



def findLastBoldRowAtBeggining(ws, minCol=1, minRow=1):
    msgText=''

    try:
        #if ws is None:
        #    ws = Worksheet
        
        lastBoldRowAtBeggining = -1
        rowTemp = minRow

        for col in ws.iter_cols(min_col=minCol, min_row=rowTemp):
          
            for cell in col:
                
                if cell.font   and   cell.font.bold:
                    lastBoldRowAtBeggining = cell.row
                    rowTemp = lastBoldRowAtBeggining
                else:
                    break
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while searching for the last bold row at the beggining.', getTraceback(e))

    if msgText: print(msgText)

    return lastBoldRowAtBeggining