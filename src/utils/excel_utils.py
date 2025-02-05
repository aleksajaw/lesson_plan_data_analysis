from error_utils import handleErrorMsg, getTraceback
#from src.constants.paths_constants import scheduleClassesExcelPath
from src.constants.conversion_constants import excelEngineName, draftSheetName
from src.constants.schedule_structures_constants import excelMargin, excelDistance
from pandas import ExcelWriter, DataFrame
from openpyxl import load_workbook#, Workbook, worksheet
from openpyxl.cell.cell import MergedCell as openpyxlMergedCell
#from openpyxl.worksheet.worksheet import Worksheet



###   DRAFTS   ###
def createDraftSheet(excelFilePath):
    msgText=''
    try:
        with ExcelWriter(excelFilePath, engine=excelEngineName, mode='w+') as writer:
            draftDf = DataFrame()  # Create an empty DataFrame
            draftDf.to_excel(writer, sheet_name=draftSheetName, startrow=excelMargin['row'], startcol=excelMargin['col'], merge_cells=True)

    except Exception as e:
        msgText = handleErrorMsg('\nError while creating draft sheet for Excel file.', getTraceback(e))

    if msgText: print(msgText)



def createDraftSheetIfNecessary(excelFilePath):
    from files_utils import doesFileExist
    if not doesFileExist(excelFilePath):
        createDraftSheet(excelFilePath)



def delDraftIfNecessary(workbook, excelFilePath):
    from files_utils import doesFileExist
    msgText=''

    if not bool(workbook) and doesFileExist(excelFilePath):

        try:
            workbook = load_workbook(excelFilePath)

            if (len(workbook.sheetnames)>1) & doesSheetExist(workbook, draftSheetName):
                deleteExcelSheet(workbook, draftSheetName)
                workbook.save(excelFilePath)
          
            else:
                workbook.close()


        except Exception as e:
            msgText = handleErrorMsg('\nUnable to open the Excel file to check and delete the draft sheet.', getTraceback(e))
    
    if msgText: print(msgText)
    


###   SHEET OPERATIONS   ###
def doesSheetExist(workbook, sheetName):
    #if workbook is None:
    #    workbook = Workbook
    return bool(sheetName in workbook.sheetnames and len(workbook.sheetnames)>0)



def deleteExcelSheet(workbook, sheetName):
    msgText = ''

    try:
        #if workbook is None:
        #    workbook = Workbook
        
        worksheet = workbook[sheetName]
        workbook.remove(worksheet)
        msgText = f'The sheet {sheetName} deleted.'
        
    except Exception as e:
        msgText = handleErrorMsg(f'\nError deleting the sheet {sheetName}.', getTraceback(e))

    if msgText: print(msgText)



def get1stNotMergedCell(group):
    foundNotMergedCell = False
    i = -1

    while not foundNotMergedCell:
        i+=1
        if not isinstance(group[i], openpyxlMergedCell):
            foundNotMergedCell = True

    return group[i] if i>=0 else None



def getNrOfLastNonEmptyCellInCol(ws, minRow=0, col=0):
    msgText=''
    counter = -1

    try:
        #if ws is None:
        #    ws = Worksheet
        
        for row in ws.iter_rows(min_row=minRow, min_col=col, max_col=col):

            if row[0].value is not None:           
                if counter < 0:
                    counter = 0
                
                counter += 1
            
            else:
                break
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while getting the last non-empty cell in column.', getTraceback(e))
                    
    if msgText: print(msgText)

    return counter

                    
def removeLastEmptyRowsInDataFrames(elToBeFiltered):
    msgText = ''

    try:
        # keep all the rows up to the last non-empty row
        for singleWorksheet in elToBeFiltered:
            if len(singleWorksheet.items()):
                for sheetName, sheetVal in singleWorksheet.items():
                    #lastNonEmptyRow = int(sheetVal.dropna(how='all').index[-1][0])
                    singleWorksheet[sheetName] = sheetVal.loc[:sheetVal.last_valid_index()]
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while removing last empty rows in Excel worksheet.', getTraceback(e))
    
    if msgText: print(msgText)

    #return elToBeFiltered



def dropnaInDfByAxis(el, axis=-1, both=True):
    msgText=''

    try:
        #if el is None:
        #    el = DataFrame
        
        axisList = [axis   if (axis>=0 and not both)   else 0,1]
        for axis in axisList:
            el = el.dropna(axis=axis, how='all')

    except Exception as e:
        msgText = handleErrorMsg('\nError while dropping the NA values in the both axis of Data Frame.', getTraceback(e))

    if msgText: print(msgText)
 
    return el



def countInnerCoords(df, writingDirection='row', innerCoords={'row':0,'col':0}):
    #if df is None:
    #    df = DataFrame
    
    if   writingDirection == 'row':
        innerCoords['col'] = innerCoords['col'] + df.shape[1] + df.index.nlevels + excelDistance['col']

    elif writingDirection == 'col':
        innerCoords['row'] = innerCoords['row'] + df.shape[0] + df.columns.nlevels + excelDistance['row']

    return innerCoords