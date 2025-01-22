from src.utils.error_utils import handleErrorMsg, getTraceback
from src.constants import scheduleExcelClassesPath, excelEngineName, draftSheetName, dfColNamesTuples, timeIndexNames
import json
import re
from pandas import ExcelWriter, DataFrame, read_excel, MultiIndex
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment as openpyxlAlignment
from openpyxl.styles import PatternFill as openpyxlPatternFill
from openpyxl.styles import Border as openpyxlBorder
from openpyxl.styles import Side as openpyxlSide
from openpyxl.cell.cell import Cell as openpyxlCell
from openpyxl.cell.cell import MergedCell as openpyxlMergedCell
from openpyxl.utils import column_index_from_string
import os



###   DRAFTS   ###
def createDraftSheet(excelFilePath=scheduleExcelClassesPath):
    try:
        with ExcelWriter(excelFilePath, engine=excelEngineName, mode='w+') as writer:
            draftDf = DataFrame()  # Create an empty DataFrame
            draftDf.to_excel(writer, sheet_name=draftSheetName, merge_cells=True)

    except Exception as e:
        print(handleErrorMsg('\nError while creating draft sheet for Excel file.', getTraceback(e)))



def createDraftSheetIfNecessary(excelPath=scheduleExcelClassesPath):
    from files_utils import doesFileExist
    if not doesFileExist(excelPath):
        createDraftSheet()



def delDraftIfNecessary(workbook=Workbook(), excelFilePath=scheduleExcelClassesPath):
    from files_utils import doesFileExist

    if not bool(workbook) and doesFileExist(excelFilePath):

        try:
            workbook = load_workbook(excelFilePath)

            if (len(workbook.sheetnames)>1) & doesSheetExist(workbook, draftSheetName):
                deleteExcelSheet(workbook, draftSheetName)
                workbook.save(excelFilePath)
          
            else:
                workbook.close()


        except Exception as e:
            print(f"Unable to open the Excel file to check and delete the draft sheet: {getTraceback(e)}")
            return

    


###   SHEET OPERATIONS   ###
def doesSheetExist(workbook=Workbook(), sheetName=''):
    return bool(sheetName in workbook.sheetnames and len(workbook.sheetnames)>0)



def deleteExcelSheet(workbook=Workbook(), sheetName=''):
    msgText = ''

    try:
        if isinstance(workbook, Workbook):
            worksheet = workbook[sheetName]
            workbook.remove(worksheet)
            msgText = f'The sheet {sheetName} deleted.'

        else:
          raise Exception(f'workbook variable should be Workbook() type, not {type(workbook)}')
        
    except Exception as e:
        msgText = handleErrorMsg(f'\nError deleting the sheet {sheetName}.', getTraceback(e))

    if msgText: print(msgText)



def writeAsDfToExcelSheet(desire=None, sheetName='', dataToEnter=None):
    msgText = ''

    # desire should be Excel.Writer or filePath
    if not desire:
        from files_utils import createFileNameWithNr
        desire = createFileNameWithNr

    try:
        df = convertToDf(dataToEnter)
        df.to_excel(desire, sheet_name=sheetName, merge_cells=True)
        msgText = f'\nData for sheet {sheetName} loaded.'

    except Exception as e:
        msgText = handleErrorMsg(f'\nError loading data into {sheetName}.', getTraceback(e))

    if msgText: print(msgText)



def writeObjOfDfsToExcel(writer=ExcelWriter, scheduleExcelClassesPath='', dataToEnter=None, isConverted=True):
    msgText = ''

    try:
        # group contains classes, teachers, subjects
        groupDfs = convertToObjOfDfs(dataToEnter)   if not isConverted   else dataToEnter

        for groupName in groupDfs:   
            groupDfs[groupName].to_excel(writer, sheet_name=delInvalidChars(groupName), merge_cells=True)

        msgText = '\nData loaded into the schedule Excel file: ' + os.path.basename(scheduleExcelClassesPath)

    except Exception as e:
        msgText = handleErrorMsg('\nError loading complete classes data.', getTraceback(e))
    
    if msgText: print(msgText)



def writeSortedObjOfDfsToExcel(objOfDfs=None, titleForDisplay='', excelPath=''):
    from schedule_utils import autoFormatScheduleExcel
    msgText = ''

    try:
        #if len(objOfDfs.keys()):
        sortedObjOfDfs = {key: objOfDfs[key] for key in sorted(objOfDfs)}
        with ExcelWriter(excelPath, mode='w+', engine=excelEngineName) as writer:       
            writeObjOfDfsToExcel(writer, excelPath, sortedObjOfDfs)
            autoFormatScheduleExcel(writer.book, excelPath)
                
    except Exception as e:
        msgText = handleErrorMsg(f'\nError while writing to the {titleForDisplay}\' Excel file.', getTraceback(e))
    
    if msgText: print(msgText)



def autoFormatExcelCellSizes(workbook=None, excelFilePath=scheduleExcelClassesPath):

    msgText = ''
    
    try:
        if not isinstance(workbook, Workbook):
            workbook = load_workbook(excelFilePath)

        if (workbook):
            rowsCounter = 0

            for ws in workbook.worksheets:
                rowsCounter+=1
                rowsLines = {}
                colsLength = {}


                for col in ws.columns:

                    colLetter = get1stNotMergedCell(col).column_letter
                    colIndex = column_index_from_string(colLetter)
                    
                    # init content length for specific column in worksheet
                    colsLength[colIndex] = 1


                    for cell in col:

                        if cell.row not in rowsLines:
                            rowsLines[cell.row] = 1

                        maxRowLines = rowsLines[cell.row]
                        maxColLength = colsLength[cell.column]
                        
                        if isinstance(cell.value, str) and '\n' in cell.value:

                            linesCount = cell.value.count('\n') + 1
                            rowsLines[cell.row] = max(maxRowLines, linesCount)
                            temp = cell.value.split('\n')

                            for t in temp:
                                colsLength[cell.column] = max(maxColLength, len(str(t)))

                        else:
                            colsLength[cell.column] = max(maxColLength, len(str(cell.value)))

                        cell.alignment = openpyxlAlignment(wrap_text=True, horizontal='center', vertical='center')


                    ws.column_dimensions[colLetter].width = colsLength[colIndex] + 2

                defaultFontSize = 11
                for row in range(1,len(rowsLines)+1):
                    ws.row_dimensions[row].height = rowsLines[row] * defaultFontSize * 1.3

            workbook.save(excelFilePath)

    except Exception as e:
        msgText = handleErrorMsg('\nError while formatting the cell sizes in the Excel file.', getTraceback(e))

    if msgText: print(msgText)



def formatCellBorder(cell=None, right='', left='', top='', bottom=''):
    msgText = ''

    if isinstance(cell, (openpyxlCell, openpyxlMergedCell)):
        try:
            currentBorder = cell.border
            defaultColor = '000000'
            borderStyle = { 'hair':   openpyxlSide(border_style='hair',   color=defaultColor),
                            'thin':  openpyxlSide(border_style='thin',  color=defaultColor),
                            'medium': openpyxlSide(border_style='medium', color=defaultColor),
                            'thick':   openpyxlSide(border_style='thick',   color=defaultColor) }
            
            cell.border = openpyxlBorder( right = borderStyle[right]    if right    else currentBorder.right,
                                          left = borderStyle[left]      if left     else currentBorder.left,
                                          top = borderStyle[top]        if top      else currentBorder.top,
                                          bottom = borderStyle[bottom]  if bottom   else currentBorder.bottom )

        except Exception as e:
            msgText = handleErrorMsg('\nError while formatting the cell.', getTraceback(e))
            
    else:
        msgText = "\nError while formatting the cell: The value must be of type 'Cell'."
    
    if msgText: print(msgText)



def formatCellBackground(cell=None, fillType='', startColor='', endColor=''):
    msgText = ''

    if isinstance(cell, (openpyxlCell, openpyxlMergedCell)):
        try:
            if (  fillType=='solid'
                  and isinstance(startColor, str) and startColor != ''
                  and isinstance(startColor, str) and endColor != '' ):
                
                cell.fill = openpyxlPatternFill(start_color=startColor, end_color=endColor, fill_type=fillType)
            
            elif fillType:
                cell.fill = openpyxlPatternFill(fill_type=fillType)

        except Exception as e:
            msgText = handleErrorMsg('\nError while formatting the cell.', getTraceback(e))
            
    else:
        msgText = '\nError while formatting the cell: The value must be of type \'Cell\'.'

    if msgText: print(msgText)



def addBgToExcelSheetRowsBasedOnObj(writer=ExcelWriter, sheetRowsToColor={'rows':[], 'colsLength':0}):
    # add BACKGROUND to the (odd here) groups of the cells in worksheet 
    msgText = ''

    try:
        workbook = writer.book
        if workbook:
                            
            for sheetname in workbook.sheetnames:
                ws = workbook[sheetname]
                sheetBgRanges = sheetRowsToColor[sheetname]

                for grRow in sheetBgRanges['rows']:
                    for col in range(1, sheetBgRanges['colsLength']+1):
                        cell = ws.cell(row=grRow, column=col)
                        formatCellBackground(cell, 'solid', 'f3f3f3', 'f3f3f3')
    
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while adding background to the cells in the Excel sheet rows.', getTraceback(e))
    
    if msgText: print(msgText)



def get1stNotMergedCell(group=[]):
    foundNotMergedCell = False
    i = -1

    while not foundNotMergedCell:
        i+=1
        if not isinstance(group[i], openpyxlMergedCell):
            foundNotMergedCell = True

    return group[i] if i>=0 else None



def getNrOfLastNonEmptyCellInCol(ws=None, minRow=int, col=int):
    msgText=''
    counter = -1

    try:
        for row in ws.iter_rows(min_row=minRow, min_col=col, max_col=col):

            if row[0].value is not None:           
                if counter < 0:
                    counter = 0
                
                counter += 1
            
            else:
                break
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while getting the last non-empty cell in column.', getTraceback(e))
                    
    if msgText: print(msgText)

    return counter

                    
def removeLastEmptyRowsInDataFrames(elToBeFiltered=None):
    msgText = ''

    try:
        if not isinstance(elToBeFiltered, list):
            elToBeFiltered = list(elToBeFiltered)

        # keep all the rows up to the last non-empty row
        for singleWorksheet in elToBeFiltered:
            if len(singleWorksheet.items()):
                for sheetName, sheetVal in singleWorksheet.items():
                    lastNonEmptyRow = int(sheetVal.dropna(how='all').index[-1][0])
                    #lastNonEmptyRow = (int(lastNonEmptyRow[0]), lastNonEmptyRow[1])
                    singleWorksheet[sheetName] = sheetVal.loc[:lastNonEmptyRow]
    
    except Exception as e:
        msgText = handleErrorMsg('\nError while removing last empty rows in Excel worksheet.', getTraceback(e))
    
    if msgText: print(msgText)

    #return elToBeFiltered



def dropnaInDfByAxis(el=None, axis=-1, both=True):
    msgText=''

    try:
        if isinstance(el, DataFrame):
            axisList = [axis   if (axis>=0 and not both)   else 0,1]
            for axis in axisList:
                el = el.dropna(axis=axis, how='all')

    except Exception as e:
        msgText = handleErrorMsg('\nError while dropping the NA values in the both axis of Data Frame.', getTraceback(e))

    if msgText: print(msgText)
 
    return el



###    CONVERTERS    ###


# DATA   =>   DATA FRAME
def convertToDf(dataToConvert=None, rowIndexesAsList=timeIndexNames, colNamesAsTuples=dfColNamesTuples):
    df = None
    msgText=''

    try:
        if(dataToConvert):

            # multi-dimensional column names
            lessonColumns = MultiIndex.from_tuples(tuples = colNamesAsTuples)

            # old columns version
            #lessonColumns = dataToConvert[0]

            # schedule without column names
            lessonRows = dataToConvert[1:]

            df = DataFrame(data=lessonRows, columns=lessonColumns)

            # use empty string instead of null/NaN
            df = df.fillna('')
            
            # Restore 111 from '111.0'.
            # The problem is probably caused by the creation of the DataFrame.
            df = df.map(convertFloatToInt)

            # Useful if there are repeated index cells in the table,
            # e.g. when there are more than one lesson
            # at the same time for one class and its groups.

            # FOR NOW, LEAVE THIS AS A COMMENT
            # IF YOU WANT TO KEEP CREATING THE TEACHERS' TIMETABLE FUNCTIONAL.
            #for indexName in timeIndexNames:
            #    df[indexName] = df[indexName].where(df[indexName] != df[indexName].shift(), '')
            
            # set actual columns as row indices
            df.set_index(keys=rowIndexesAsList, inplace=True)
            
    except Exception as e:
        msgText = handleErrorMsg('\nError while converting data do DataFrame.', getTraceback(e))

    if msgText: print(msgText)  
    
    return df



# DATA   =>   OBJECT OF DATA FRAMES 
def convertToObjOfDfs(dataToConvert=None):
    if dataToConvert:
        return {sheetName: convertToDf(dataToConvert[sheetName])   for sheetName in dataToConvert}

    else:
        return {draftSheetName: DataFrame()}



# OBJECT OF DATA FRAMES   =>   JSON
def convertObjOfDfsToJSON(dataToConvert=None):
    msgText = ''
    objOfDfsJSON = {}

    try:
        for sheetName, df in dataToConvert.items():
            objOfDfsJSON[sheetName] = df.to_json(orient='split')

        objOfDfsJSON = json.dumps(objOfDfsJSON, indent=4)

    except Exception as e:
        msgText = handleErrorMsg('\nError while converting object of DataFrames to JSON.', getTraceback(e))

    if msgText: print(msgText)
    return objOfDfsJSON



# EXCEL CONTENT   =>   OBJECT OF DATA FRAMES
#                    =>   JSON
def convertExcelToDfsJSON(defaultIndexes = timeIndexNames):
    from files_utils import doesFileExist
    excelJSON = {}
    msgText = ''

    if not doesFileExist(scheduleExcelClassesPath):
        return excelJSON

    try:
        excelData = read_excel(io=scheduleExcelClassesPath, sheet_name=None, engine=excelEngineName,
                                keep_default_na=False)

        if(bool(excelData)):

            # old columns version
            #lessonColumns = dataToConvert[0]

            for sheetName, oldDf in excelData.items():
            
                if not oldDf.empty:
                    df = DataFrame(oldDf[2:])

                    # multi-dimensional column names
                    df.columns = MultiIndex.from_tuples(tuples = dfColNamesTuples)

                    # useful if there are repeated index cells in the table
                    # e.g. when there are more than one lesson
                    # at the same time for one class and its groups
                    for indexName in defaultIndexes:
                        df[indexName] = df[indexName].where(df[indexName] != df[indexName].shift(), '')
                    
                    df.set_index(keys=defaultIndexes, inplace=True)
                
                else:
                    df = oldDf
                excelData[sheetName] = df.to_json(orient='split')

        excelJSON = json.dumps(excelData, indent=4)


    except Exception as e:
        msgText = handleErrorMsg('\nError converting existing schedule Excel file to JSON.', getTraceback(e))

    if msgText: print(msgText)


    return excelJSON 



# 111.0   =>   111
def convertFloatToInt(value=None):
    isValueFloat = isinstance(value, float) and value.is_integer()
    if isValueFloat:
        return int(value)
    return value


# '1'   =>   1
def convertDigitInStrToInt(text=''):
    return int(text)  if str.isdigit(text)  else text


# <br>
# <br />   =>   \n
def convertBrInText(text=''):
    if isinstance(text, str):
        text = text.replace("<br>", "\n").replace("<br />", "\n")

    return text



# <p>
#   <span>hello</span> <span>world!&nbsp;</span>
# </p>
#   =>   helloworld!
def splitHTMLAndRemoveTags(HTMLText=''):
    #patternSub = r'&nbsp;|\s+'
    patternSub = r'&nbsp;'
    HTMLTextStripped = re.sub(patternSub, '', HTMLText)
    textParts = []
    convertedTextParts = []

    if HTMLTextStripped:
        # pattern to get e.g. string 'text' from <span class='p'>text</span>
        pattern = r'<[^>]+>([^<]+)</[^>]+>'
        textParts = re.findall(pattern, HTMLTextStripped)   or   [HTMLTextStripped]
        convertedTextParts = [convertDigitInStrToInt(part)   for part in textParts]      

    return convertedTextParts



# 'part 1/2'    =>   'part 1_2'
# or '[part1]'   =>   '_part1_'
def delInvalidChars(name='', target='sheetName'):
    if target=='sheetName':
        invalidScheetNameChars = ['/', '\\', ':', '*', '?', '[', ']']
        # replace invalid characters with character '_'
        return ''.join('_'  if c in invalidScheetNameChars   else c   for c in name)
    else:
        return name
    


# for example,
# {'key1':[], 'key2':[]}   =>   ['key1', 'key2']
def getListOfKeys(obj={}):
    return list(obj.keys())



def filterNumpyNdarray(arr=np.ndarray, elToDel=''):
    # convert values to string
    arrAsStr = arr.astype(str)
    # remove specific value
    sortedArr = np.sort( arrAsStr[ arrAsStr != elToDel] )

    return np.array([val for val in sortedArr])