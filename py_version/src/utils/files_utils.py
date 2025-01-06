import os
import sys
import subprocess
from constants import scheduleExcelPath, outputsPath
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment as openpyxlAlignment
from openpyxl.styles import PatternFill as openpyxlPatternFill
from openpyxl.styles import Border as openpyxlBorder
from openpyxl.styles import Side as openpyxlSide
from openpyxl.cell.cell import Cell as openpyxlCell
from openpyxl.cell.cell import MergedCell as openpyxlMergedCell
from openpyxl.utils import column_index_from_string, get_column_letter



def doesFileExist(filePath='', shouldPrintMsg=False):
    msgText = f'File   {os.path.basename(filePath)}   '
    doesFileExistBool = bool (os.path.isfile(filePath))

    if doesFileExistBool:
        msgText += 'exists.'
    else:
        msgText += 'does not exist.'

    if shouldPrintMsg:
        print(msgText)

    return doesFileExistBool



def compareAndUpdateFile(filePath='', dataToCompare=''):
    
    isFileChanged = False
    if bool(filePath) and bool(dataToCompare):
        msgText = ''

        try:
            with open(file=filePath, mode="r+") as file:
                #file.seek(0)
                fileContent = file.read()
                if str(fileContent) != str(dataToCompare):
                    file.seek(0)
                    file.write(dataToCompare)
                    # make sure to delete old redundant value
                    file.truncate()
                    msgText = f'\nFile   {os.path.basename(filePath)}   updated with new data.'
                    isFileChanged = True
                    
                else:
                    msgText = f'\nNothing to be updated in file   {os.path.basename(filePath)}.'

                file.close()
                print(msgText)

        except FileNotFoundError as e:
            
            with open(filePath, 'w') as file:
              file.write(dataToCompare)
              file.close()

              print(f'File   {os.path.basename(filePath)}   not found. Created a new file and complete it with data.')
        
        except Exception as e:
            print('Error while comparing and updating file content: ', e)


        return isFileChanged



def autoFormatMainExcelFile(workbook=Workbook(), excelFilePath=scheduleExcelPath):
    autoFormatExcelFileCellSizes(workbook, excelFilePath)
    autoFormatMainExcelFileCellsStyle(workbook, excelFilePath)



def autoFormatExcelFileCellSizes(workbook=None, excelFilePath=scheduleExcelPath):
    from excel_utils import get1stNotMergedCell

    try:
        if not isinstance(workbook, Workbook):
            workbook = load_workbook(excelFilePath)

        if (workbook):
            rowsCounter = 0

            for ws in workbook.worksheets:
                rowsCounter+=1
                rowsLines = {}
                colsLength = {}


                for col in ws.columns:

                    colLetter = get1stNotMergedCell(col).column_letter
                    colIndex = column_index_from_string(colLetter)
                    
                    # init content length for specific column in worksheet
                    colsLength[colIndex] = 1


                    for cell in col:

                        if cell.row not in rowsLines:
                            rowsLines[cell.row] = 1

                        maxRowLines = rowsLines[cell.row]
                        maxColLength = colsLength[cell.column]
                        
                        if isinstance(cell.value, str) and '\n' in cell.value:

                            linesCount = cell.value.count('\n') + 1
                            rowsLines[cell.row] = max(maxRowLines, linesCount)
                            temp = cell.value.split('\n')

                            for t in temp:
                                colsLength[cell.column] = max(maxColLength, len(str(t)))

                        else:
                            colsLength[cell.column] = max(maxColLength, len(str(cell.value)))

                        cell.alignment = openpyxlAlignment(wrap_text=True, horizontal='center', vertical='center')


                    ws.column_dimensions[colLetter].width = colsLength[colIndex] + 2

                defaultFontSize = 11
                for row in range(1,len(rowsLines)+1):
                    ws.row_dimensions[row].height = rowsLines[row] * defaultFontSize * 1.3

            workbook.save(excelFilePath)

    except Exception as e:
        print('Error while formatting the Excel file:', e)



def formatCellBorder(cell=None, right='', left='', top='', bottom=''):
    if isinstance(cell, (openpyxlCell, openpyxlMergedCell)):
        try:
            currentBorder = cell.border
            defaultColor = '000000'
            borderStyle = { 'hair':   openpyxlSide(border_style='hair',   color=defaultColor),
                            'thin':  openpyxlSide(border_style='thin',  color=defaultColor),
                            'medium': openpyxlSide(border_style='medium', color=defaultColor),
                            'thick':   openpyxlSide(border_style='thick',   color=defaultColor) }
            
            cell.border = openpyxlBorder( right = borderStyle[right]    if right    else currentBorder.right,
                                          left = borderStyle[left]      if left     else currentBorder.left,
                                          top = borderStyle[top]        if top      else currentBorder.top,
                                          bottom = borderStyle[bottom]  if bottom   else currentBorder.bottom )

        except Exception as e:
            print('Error while formatting the cell:', e)
            
    else:
        print("Error while formatting the cell: The value must be of type 'Cell'.")



def formatCellBackground(cell=None, fillType='', startColor='', endColor=''):
    if isinstance(cell, (openpyxlCell, openpyxlMergedCell)):
        try:
            if (  fillType=='solid'
                  and isinstance(startColor, str) and startColor != ''
                  and isinstance(startColor, str) and endColor != '' ):
                cell.fill = openpyxlPatternFill(start_color=startColor, end_color=endColor, fill_type=fillType)
            
            elif fillType:
                cell.fill = openpyxlPatternFill(fill_type=fillType)

        except Exception as e:
            print('Error while formatting the cell:', e)
            
    else:
        print("Error while formatting the cell: The value must be of type 'Cell'.")


def autoFormatMainExcelFileCellsStyle(workbook=Workbook(), excelFilePath=scheduleExcelPath):
    from constants import weekdays, lessonAttrs

    try:
        if not isinstance(workbook, Workbook):
            workbook = load_workbook(excelFilePath)

        if (workbook):
            for ws in workbook.worksheets:
                
                permEmptyCellStyle = openpyxlPatternFill(fill_type='lightTrellis')
                
                # merge and format empty cells in the corner between the MultiIndexes
                ws.merge_cells('A1:B2')
                ws['A1'].fill = permEmptyCellStyle

                rowsStart = 3
                rowsCount = 0

                # count rows in the 1st and the 2nd column,
                # cells there are less likely to be merged (for easier data analysis)
                for row in ws.iter_rows(min_row=rowsStart, min_col=1, max_col=2):
                    if row[1].value is not None:
                        rowsCount += 1
                    else:
                        break
                

                totalRowsCount = rowsStart-1 + rowsCount
                rowIndexesColsEnd = 2
                colIndexesRowsEnd = 2
                colDaysStart = 3
                lenOfWeekdays = len(weekdays)
                lenOfLessonAttrs = len(lessonAttrs)
                totalColsCount = colDaysStart-1 + lenOfWeekdays * lenOfLessonAttrs
                

                # row with number rowsStart in the 1st two columns
                # contains the names for the rows' MultiIndex
                allEmpty = True 

                # check if the cells in the 1st row below column indexes are empty
                for col in range(colDaysStart, totalColsCount+1):
                    
                    if ws.cell(row=rowsStart, column=col).value is not None:
                        allEmpty = False
                        break


                # change the BACKGROUND of the cells in the row below column indexes
                # only if entire row is empty
                if allEmpty:
                    
                    # MERGE CELLS in the row 
                    ws.merge_cells( start_row=rowsStart, start_column=colDaysStart,
                                    end_row=rowsStart,   end_column=totalColsCount )
                    
                    cell = ws.cell(row=rowsStart, column=colDaysStart)
                    cell.fill = permEmptyCellStyle

                    #startCoordinate = f'{get_column_letter(colDaysStart)}{rowsStart}'
                    #endCoordinate = f'{get_column_letter(totalColsCount)}{rowsStart}'
                    
                    #for cell in ws[startCoordinate:endCoordinate][0]:
                    #    formatCellBorder(cell, bottom='thin')


                # add MEDIUM RIGHT BORDERS
                # at the end of each important column and each day
                dayColsEnd = [colIndexesRowsEnd + i * 3 for i in range(1, 6)]
                rowIndexesColsList = list(range(1, colIndexesRowsEnd+1))
                thickerColsList = rowIndexesColsList + dayColsEnd

                for thickerCol in thickerColsList:
                    for row in range(1, totalRowsCount+1):
                        cell = ws.cell(row=row, column=thickerCol)
                        formatCellBorder(cell, right='medium')
                        # add THIN SIDE BORDERS to the center column on days
                        if thickerCol > rowIndexesColsEnd:
                            cell = ws.cell(row=row, column=thickerCol-1)
                            formatCellBorder(cell, left='thin', right='thin')
                

                # add MEDIUM BOTTOM BORDERS
                # at the end of each index and the entire schedule
                colIndexesRowsList = list(range(1, colIndexesRowsEnd+1))[1:]
                thickerRowsList = colIndexesRowsList + [rowsStart, totalRowsCount]

                for thickerRow in thickerRowsList:
                    colsLimit = totalColsCount+1
                    # ignore long merged cell
                    if thickerRow==rowsStart:
                        colsLimit = rowIndexesColsEnd+1

                    for col in range(1, colsLimit):
                        cell = ws.cell(row=thickerRow, column=col)
                        formatCellBorder(cell, bottom='medium')


                # add BACKGROUND for the lessons with an odd numbers
                # and HAIR/THIN BORDERS
                #rowsToBeColoured = []
                for row in range(1, ws.max_row + 1):
                    # fast checking if the row is in the specific merged range (for col A here)
                    merged = next(  ( r   for r in ws.merged_cells.ranges
                                          if r.min_col == 1 and r.min_row <= row <= r.max_row ),
                                    None)
                

                    for col in range(1, totalColsCount+1):
                        cellForBorder = ws.cell(row=row, column=col)
                        # thinner ('HAIR') BORDER inside merged row
                        #if merged and merged.min_row != row:
                        #    topBorderStyle = 'hair'
                        # THIN BORDER for outer frame
                        #elif not cellForBorder.border.top.style:
                        topBorderStyle = 'thin'
                        formatCellBorder(cellForBorder, top=topBorderStyle)


                    # set LIGHT GREY BACKGROUND
                    rowNr = merged.min_row   if merged   else row
                    cellValue = ws.cell(row=rowNr, column=1).value
                    if isinstance(cellValue, int) and (cellValue % 2 != 0):
                        #rowsToBeColoured.append(row)
                        for col in range(1, totalColsCount+1):
                            cell = ws.cell(row=row, column=col)
                            formatCellBackground(cell, fillType='solid', startColor='E5E5E5', endColor='E5E5E5')


    except Exception as e:
        print('Error while formatting the Excel file:', e)



def findFileGroup(desiredBase='', desiredExt='', getSplitFileName=False):
    from collections import defaultdict
    group = defaultdict(list)

    for fileName in os.listdir(outputsPath):
        basename, ext = splitFileName(fileName)

        if desiredExt in ('', '*', ext) and desiredBase in ('', '*', basename):
            el = (basename, ext)   if getSplitFileName   else fileName

            group.append(el)

    return group



def splitFileName(fileName=''):
    basename, ext = ('', '')

    if fileName:
        basename, ext = os.path.splitext(fileName)

    return (basename, ext)



def findLastFileInGroup(desiredBase='', desiredExt='', getSplitFileName=False):
    filesList = findFileGroup(desiredBase, desiredExt, getSplitFileName)

    return filesList[-1]   if len(filesList)   else None



def getFileMarker(fileName='', separator='-'):
    if separator in fileName:
        return removeEmptyStrFromArr(fileName.split(separator))[-1]

    else:
        return fileName



def createFileName(basicFileName = 'schedule', fileExt = 'xlsx', separator = '-'):
    fileNameParts = findLastFileInGroup(basicFileName, fileExt, True)
    difference = ''

    # If there aren't any very similar files in the folder,
    # use the basic file name.
    finalFileName = basicFileName

    # Otherwise, expand the basic file name.
    if len(fileNameParts):
        difference = findFileNameDifference(fileNameParts[0])

        # But still - only if it is very similar file name.
        # For example basicFileName + separator + 1.fileExt
        hasSeparator = separator in difference
        if hasSeparator:
            tempDifference = getFileMarker(difference, separator)
            
            if len(tempDifference) == 1 and type(tempDifference) == int:
                difference = tempDifference+1
        
        else:
            difference = 1

        finalFileName+= separator + difference


    return finalFileName + '.' + fileExt

        

def findFileNameDifference(fileName='', fileNameToBeCompared=''):
    from excel_utils import convertDigitInStrToInt

    difference = ''
    nameToRemove=''
    isReplacementRequired = False

    if fileNameToBeCompared in fileName:
        isReplacementRequired = True


    difference = fileName   if not isReplacementRequired   else fileNameToBeCompared
    nameToRemove = fileNameToBeCompared   if not isReplacementRequired   else fileName

    difference.replace(nameToRemove, '')

    return convertDigitInStrToInt(difference)



def removeEmptyStrFromArr(arr=[]):
    return [el   for el in arr   if el!='']



def openFileWithDefApp(filePath=''):

    if filePath!='' and not doesFileExist(filePath, True):
        return

    try:
        # Windows
        if sys.platform == "win32":
            subprocess.run(['start', '', filePath], shell=True)

        # macOS
        elif sys.platform == "darwin":
            subprocess.run(['open', filePath])

        # Linux / Unix-like
        else:
            subprocess.run(['xdg-open', file_path])

    except Exception as e:
        print(f"Failed to open file for User: {e}")