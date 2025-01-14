from src.utils.error_utils import getTraceback
from src.constants import scheduleExcelClassesPath, weekdays, timeIndexNames, dfColWeekDayNamesTuples3el, dfColWeekDayNamesTuples4el, lessonTimePeriods, dfColWeekDayEmptyRow
from src.utils import autoFormatExcelCellSizes, formatCellBackground, formatCellBorder, dropnaInDfByAxis
import pandas as pd
from pandas import DataFrame
import numpy as np
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill as openpyxlPatternFill



def autoFormatScheduleExcel(workbook=Workbook(), excelFilePath=scheduleExcelClassesPath):
    autoFormatExcelCellSizes(workbook, excelFilePath)
    autoFormatScheduleExcelCellStyles(workbook, excelFilePath)



def mergeEmptyCellsAndColorBg(ws=None, mergedCellObj={'startRow':int, 'startCol': int, 'endRow': int, 'endCol': int}):
    msgText=''

    try:
        permEmptyCellStyle = openpyxlPatternFill(fill_type='lightTrellis')

        ws.merge_cells( start_row=mergedCellObj['startRow'],
                        start_column=mergedCellObj['startCol'],
                        end_row=mergedCellObj['endRow'],
                        end_column=mergedCellObj['endCol'] )

        ws.cell(row=mergedCellObj['startRow'], column=mergedCellObj['startCol']).fill = permEmptyCellStyle


    except Exception as e:
        msgText = f'\nError while merging the empty cells and coloring their background: {getTraceback(e)}'


    if msgText: print(msgText)



def colorBgOfEmptyRow(ws=None, colRange=None, row=int, startColumn=int):
    msgText=''

    try:
        permEmptyCellStyle = openpyxlPatternFill(fill_type='lightTrellis')
        allEmpty = True 

        # check if the cells in the 1st row (below column indices) are empty
        for col in colRange:
            
            if ws.cell(row=row, column=col).value is not None:
                allEmpty = False
                break


        # change the BACKGROUND of the cells in the row (below column indices)
        # only if entire row is empty
        if allEmpty:
            
            # MERGE CELLS in the row 
            ws.merge_cells( start_row=row, start_column=startColumn,
                            end_row=row,   end_column=max(colRange) )
            
            cell = ws.cell(row=row, column=startColumn)
            cell.fill = permEmptyCellStyle
            

    except Exception as e:
        msgText = f'\nError while coloring the background of empty rows: {getTraceback(e)}'

    if msgText: print(msgText)



def findLastBoldRowAtBeggining(ws, minCol=1):
    msgText=''

    try:
        lastBoldRowAtBeggining = -1
        rowTemp = 1
        for col in ws.iter_cols(min_col=minCol, min_row=rowTemp):
            for cell in col:
                if cell.font and cell.font.bold:
                    lastBoldRowAtBeggining = cell.row
                    rowTemp = lastBoldRowAtBeggining
                else:
                    break
    
    except Exception as e:
        msgText = f'\nError while searching for the last bold row at the beggining: {getTraceback(e)}'

    if msgText: print(msgText)

    return lastBoldRowAtBeggining
                        


def autoFormatScheduleExcelCellStyles(workbook=Workbook(), excelFilePath=scheduleExcelClassesPath):
    from excel_utils import getNrOfLastNonEmptyCellInCol
    msgText=''

    try:
        if not isinstance(workbook, Workbook):
            workbook = load_workbook(excelFilePath)

        if (workbook):
            for ws in workbook.worksheets:
                
                # column nr where the row indices end
                rowIndexesLastCol = 2

                # bold rows are for headers
                lastBoldRowAtBeggining = findLastBoldRowAtBeggining(ws, rowIndexesLastCol)

                # merge and format empty cells in the corner between the MultiIndexes
                mergeEmptyCellsAndColorBg(ws, { 'startRow': 1,
                                                'startCol': 1,
                                                'endRow': lastBoldRowAtBeggining,
                                                'endCol': rowIndexesLastCol } )

                lastMergedCellColIn1stRow = max([cell.max_col   for cell in ws.merged_cells
                                                                if cell.min_row==1])
                
                # cells in the 1st two columns which row nr equals contentRowsStart
                # contains the names for the rows' MultiIndex
                contentFirstRow = lastBoldRowAtBeggining+1
                # col nr for the 1st cell for the 1st day
                daysFirstColList = rowIndexesLastCol+1

                totalColsCount = lastMergedCellColIn1stRow
                totalColsRange = range(1, totalColsCount+1)
                # cells in the 2nd column contain part of rows indices & are less likely to be merged (for easier data analysis)
                totalRowsCount = lastBoldRowAtBeggining + getNrOfLastNonEmptyCellInCol(ws, minRow=contentFirstRow, col=2)                    
                totalRowsRange = range(1, totalRowsCount+1)

                colorBgOfEmptyRow(ws, range(rowIndexesLastCol+1, totalColsCount+1), row=contentFirstRow, startColumn=daysFirstColList)


                # add MEDIUM RIGHT BORDERS
                # at the end of each important column and each day
                daysLastCol = sorted([mergedCellRange.max_col  for mergedCellRange in ws.merged_cells
                                                              if mergedCellRange.min_row==mergedCellRange.max_row==1
                                                                and mergedCellRange.min_col > 2])
                rowIndexesCols = list(range(1, rowIndexesLastCol+1))
                colsWithRightMediumBorder = rowIndexesCols + daysLastCol
                standardDaySize = daysLastCol[0] - rowIndexesLastCol

                for col in colsWithRightMediumBorder:
                    for row1 in totalRowsRange:
                        cell = ws.cell(row=row1, column=col)
                        formatCellBorder(cell, right='medium')

                        # add THIN SIDE BORDERS to the center column on days
                        # do not include the columns reserved for the rows indices
                        if rowIndexesLastCol < col:
                            for attrCol in range((col-standardDaySize)+1, col):
                                cell = ws.cell(row=row1, column=attrCol)
                                formatCellBorder(cell, left='thin', right='thin')
                

                # add MEDIUM BOTTOM BORDERS
                # at the end of each index (for columns, for rows) and the entire schedule
                rowsWithBottomBorder = [lastBoldRowAtBeggining, contentFirstRow, totalRowsCount]

                for row in rowsWithBottomBorder:
                    colsLimit = totalColsCount+1
                    # ignore long merged cell
                    if row==contentFirstRow:
                        colsLimit = rowIndexesLastCol+1

                    for col in range(1, colsLimit):
                        cell = ws.cell(row, column=col)
                        formatCellBorder(cell, bottom='medium')


                # add BACKGROUND for the lessons with an odd numbers
                # and HAIR/THIN BORDERS
                #rowsToBeColoured = []
                for row in totalRowsRange:
                    colsLimit = totalColsCount+1

                    # fast checking if the row is in the specific merged range (for col A here)
                    merged = next(  ( r   for r in ws.merged_cells.ranges
                                          if r.min_col == 1 and r.min_row <= row <= r.max_row ),
                                    None)

                    for col in range(1, colsLimit):
                        cell = ws.cell(row, column=col)
                        # thinner ('HAIR') BORDER inside merged row
                        #if merged and merged.min_row != row:
                        #    topBorderStyle = 'hair'
                        # THIN BORDER for outer frame
                        #elif not cell.border.top.style:
                        topBorderStyle = 'thin'
                        formatCellBorder(cell, top=topBorderStyle)


                    # set LIGHT GREY BACKGROUND
                    rowNr = merged.min_row   if merged   else row
                    cellValue = ws.cell(row=rowNr, column=1).value

                    if isinstance(cellValue, int) and (cellValue % 2 != 0):
                        #rowsToBeColoured.append(row)
                        for col in totalColsRange:
                            cell = ws.cell(row, column=col)
                            formatCellBackground(cell, fillType='solid', startColor='E5E5E5', endColor='E5E5E5')


    except Exception as e:
        msgText = f'\nError while formatting the cell styles in the Excel file: {getTraceback(e)}'
      
    if msgText: print(msgText)



def concatAndFilterScheduleDataFrames(el1=None, el2=None, addNewCol=False, newColName='', newColVal=''):
    msgText = ''
    newDf = None

    try:
        # two lines below prevents FutureWarning:
        #   The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.
        #   In a future version, this will no longer exclude empty or all-NA columns when determining the result dtypes.
        #   To retain the old behavior, exclude the relevant entries before the concat operation.
        #el1 = dropnaInDfByAxis(el1, 1)
        if isinstance(el1, DataFrame) and isinstance(el2, DataFrame):
            el2 = dropnaInDfByAxis(el2, 1)
            newDf = pd.concat([el1, el2]).reset_index()
            newDf.set_index(keys=timeIndexNames, inplace=True)
            newDf = newDf.sort_index(level=0)
        else:
            newDf = el1 or el2
        newDf = filterAndConvertScheduleDataFrames(newDf, addNewCol, newColName, newColVal)


    except Exception as e:
        msgText = f'\nError while concatenating and filter the schedule Data Frames: {getTraceback(e)}'

    if msgText: print(msgText)

    return newDf
        


def filterAndConvertScheduleDataFrames(df=None, addNewCol=False, newColName='', newColVal=''):
    newDfFiltered = []
    msgText = ''
    
    try:
        newDf = df.copy()
        #print('\nnewDf.index ', list(newDf.index))
        rowsFiltered = []

        prepareNewColVal = addNewCol and newColName and newColVal

        colDayNamesTuples = dfColWeekDayNamesTuples4el   if addNewCol   else dfColWeekDayNamesTuples3el
        timeKey1 = timeIndexNames[0]
        timeKey2 = timeIndexNames[1]

        # iterate through rows (time indices)
        for (lessonNr, time), row in newDf.groupby(timeIndexNames):
            #print(lessonNr)
            #print(time)
            #print(row)
            rowFrame = row
            singleRowBase = {}
            singleRowBase[timeKey1] = int(lessonNr)
            singleRowBase[timeKey2] = time

            innerRows = []

            for col in newDf.columns:
                #print(rowFrame.keys())
                #print(col)
                booleanMask = rowFrame[col] != ''
                nonEmptyValues = rowFrame[col][booleanMask].dropna().tolist()
                #print(nonEmptyValues)

                if nonEmptyValues:
                    for index, value in enumerate(nonEmptyValues):
                        
                        if len(innerRows) < len(nonEmptyValues):
                            innerRows.append(singleRowBase.copy())

                            currRowNr = int( singleRowBase[timeKey1] )


                        # add empty rows to avoid tables that do not start from row nr 1
                        if ( (       rowsFiltered   and   (int(rowsFiltered[-1][timeKey1]) < currRowNr-1) )
                            or ( not rowsFiltered   and   1 < currRowNr) ):

                            lastFilteredRowNr = ( int( rowsFiltered[-1][timeKey1] )  if len(rowsFiltered)
                                                                                           else 0 )

                            missingNrs = list( range( lastFilteredRowNr+1, currRowNr ) )
                            
                            while len(missingNrs):
                                #print([r[timeKey1]   for r in rowsFiltered])
                                #print((singleRowBase[timeKey1]))
                                #print(missingNrs)
                                desiredNr = missingNrs[0]
                                singleRowTemp = {}
                                
                                for lessonAttr in colDayNamesTuples:
                                    singleRowTemp[lessonAttr] = np.nan
                                
                                desiredPreviousTime = lessonTimePeriods[ desiredNr-1 ]
                                singleRowTemp[timeKey1] = desiredNr
                                singleRowTemp[timeKey2] = desiredPreviousTime
                                rowsFiltered.append(singleRowTemp.copy())
                                missingNrs = missingNrs[1:]   if missingNrs   else []
                                singleRowTemp = {}


                        innerRows[index][col] = value
                        #print('innerRows[index]:', innerRows[index])


                    if prepareNewColVal:
                        for index, r in enumerate(innerRows):
                            if not (col[0], newColName) in r:
                                innerRows[index][(col[0], newColName)] = (newColVal   if not newColVal.isdigit()
                                                                                      else int(newColVal))

            for r in innerRows:
                rowsFiltered.append(r.copy())


        if addNewCol or ( len(newDf.columns.get_level_values(0).unique()) < len(weekdays) ):

            columnsVal = pd.MultiIndex.from_tuples(colDayNamesTuples)
        
        else:
            columnsVal = newDf.columns

        #print('rowsFiltered indexy ', [row[timeKey1]   for row in rowsFiltered])
        newDfFiltered = pd.DataFrame(rowsFiltered)
        #print('newDfFiltered.index ', newDfFiltered.index)
        newDfFiltered = newDfFiltered.reset_index()
        newDfFiltered.set_index(keys=timeIndexNames, inplace=True)
        #print('newDfFiltered.index ', newDfFiltered.index)
        newDfFiltered = newDfFiltered.reindex(columns=columnsVal, fill_value=np.nan)

    except Exception as e:
        msgText = f'\nError while filter and convert schedule Data Frames for Excel worksheet: {getTraceback(e)}'
        
    if msgText: print(msgText)
    
    return newDfFiltered



# 'r' comes from 'rozszerzony/a' which means the subject is extendend like 'advanced English'
def createGroupsInListByPrefix(data=[], splitDelimiter = '-', replaceDelimiters = ['.r', 'r_']):
    msgText = ''

    try:
        groupList = []

        # only leave the part before the first '-' and cut '.r' and 'r_' out
        for item in data:
            itemAdded = False

            if isinstance(item, str):
                item1stEl = str(item).split(splitDelimiter)[0]

                for repDeli in replaceDelimiters:
                    itemAdded = False

                    if repDeli in item1stEl:
                        
                        # e.g., delimiter to replace + subject name   =>   'r_matematyka'
                        if item1stEl.startswith(repDeli):
                            groupList.append( item1stEl.replace(repDeli, '', 1) )
                            itemAdded = True
                            break
                        
                        # e.g., delimiter to replace + subject name   =>   'matematyka.r'
                        elif item1stEl.endswith(repDeli):
                            deliIndex = item1stEl.rfind(repDeli)

                            if deliIndex != -1:
                                groupList.append( item1stEl[:deliIndex] )
                                itemAdded = True
                                break
                
                
                if not itemAdded:
                    groupList.append(item1stEl)
            
            else:
                groupList.append(item)
        
        # group elements by names starting with the same prefix
        for i in range(2, len(groupList)):
            
            el = groupList[i]
            previousEl = groupList[i-1]
            doesElStartSameAsPrevious = el.startswith( previousEl )

            if doesElStartSameAsPrevious:
                groupList[i] = previousEl
    
    except Exception as e:
        msgText = f'\nError while creating groups in list by prefix: {getTraceback(e)}'
    
    if msgText: print(msgText)
    
    return groupList



def createGroupsInListByFirstLetter(data=[]):
    msgText = ''

    try:
        newData = []
        for item in data:
            #if any(c.isdigit()   for c in item):
            match = re.match(r'[^\d\w]+[a-zA-Z]*', item)

            if match:
              if item!=match[0]:
                # e.g. '#re4'   =>   '#re'
                newData.append(match[0])

              else:
                # e.g. '#re'   =>   '#r'
                newData.append(match[1])
            
            else:
                newData.append(item[0])

        return newData
        #return [item[0]   for item in data]
    
    except Exception as e:
        msgText = f'\nError while creating groups in list by first letter: {getTraceback(e)}'
    
    if msgText: print(msgText)



def createGroupsInListByNumbers(data=[], optionalPartInNrPrefix='0'):
    msgText=''

    try:
        groupsInList = []
        for item in data:
            itemStr = str(item)
            itemLen = len(itemStr)
            
            if any(c.isalpha()   for c in itemStr):
                item = ''.join([c   for c in itemStr if c.isalpha()])

            elif itemStr.isdigit():
                if itemLen > 1:
                    item = itemStr[0] + (itemLen-1) * '0'
              
                elif 0 < int(item) < 10:
                    item = 1
                
                item = int(item)
            
            else:
                # match e.g. '.9', '_09', '09' or '9'
                pattern = r'([^a-zA-Z0-9]*' + re.escape(optionalPartInNrPrefix) + r'*[^a-zA-Z0-9]*)\d+'
                match = re.match(pattern, itemStr)
                
                if match:
                    # '.', '_0', '0'   only if   optionalPartInNrPrefix = 0
                    item = match.group(1)
            
            groupsInList.append(item)
        
        return groupsInList
    
    except Exception as e:
        msgText = f'\nError while creating groups in list by numbers: {getTraceback(e)}'
    
    if msgText: print(msgText)



def createGroupsInListBy(groupName='', data=[]):
    result = []
    msgText = ''
    
    try:
        match groupName:
            case 'subjects':
                result = createGroupsInListByPrefix(data)
        
            case 'teachers':
                result = createGroupsInListByFirstLetter(data)

            case 'classrooms':
                result = createGroupsInListByNumbers(data)


    except Exception as e:
        msgText = f'\nError while creating groups in list by...: {getTraceback(e)}'
    
    if msgText: print(msgText)

    return result